"""
Сгенерировать n списков случайных размеров со случайными числами от 1 до 10,
создать таблицу с этими списками в качестве столбцов,
используя в качестве столбцов буквы алфавита и заполнив пустые ячейки значением 0,
и сохранить таблицу в формате xlsx.
"""

import random
import string

from openpyxl import Workbook


def generate_and_save_table(n, file_name):
    columns = list(string.ascii_uppercase[:n])

    def generate_random_list(size):
        return [random.randint(1, 10) for _ in range(size)]

    data = {column: generate_random_list(random.randint(1, 10)) for column in columns}

    workbook = Workbook()
    sheet = workbook.active

    for col_num, column in enumerate(columns, start=1):
        sheet.cell(row=1, column=col_num, value=column)

        for row_num, value in enumerate(data[column], start=2):
            sheet.cell(row=row_num, column=col_num, value=value)

    for row in sheet.iter_rows(min_row=2, min_col=1):
        for cell in row:
            if cell.value is None:
                cell.value = 0

    workbook.save(file_name)


if __name__ == '__main__':
    generate_and_save_table(5, 'file.xlsx')
